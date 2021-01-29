import openpyxl

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")

def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        if row[0].value and isinstance(row[0].value, int) and row[0].value > 1 and row[0].value <1000000:
            for cell in row:
                print(cell.value, "\t", end="")
                if cell.hyperlink and cell.hyperlink.display:
                    print(cell.hyperlink.display, "\t", end="")
                    cell.hyperlink.display = '999' + cell.hyperlink.display
            print()
    # workbook.save(path)    #


book_name_xlsx = '../excel/base.xlsx'

sheet_name_xlsx = 'Sheet1'

value3 = [["姓名", "性别", "年龄", "城市", "职业"],
          ["111", "女", "66", "石家庄", "运维工程师"],
          ["222", "男", "55", "南京", "饭店老板"],
          ["333", "女", "27", "苏州", "保安"], ]

# write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)

if __name__ == '__main__':
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
